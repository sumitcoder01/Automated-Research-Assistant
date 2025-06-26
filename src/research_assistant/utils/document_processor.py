import logging
import os
import io
import json
import csv
from pathlib import Path
from typing import Optional, Dict, List

# Import necessary libraries
import pypdf
import docx
from PIL import Image
import pytesseract
import openpyxl
from pdf2image import convert_from_path

logger = logging.getLogger(__name__)

# --- Constants ---
# Minimum characters to consider direct PDF text extraction successful
MIN_PDF_TEXT_LENGTH = 100
# Supported file extensions for each processing type
TEXT_EXTENSIONS = {".txt", ".sent", ".md", ".rtf"}
JSON_EXTENSIONS = {".json"}
CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".gif"}
DOCX_EXTENSIONS = {".docx"}
PDF_EXTENSIONS = {".pdf"}


class DocumentProcessor:
    """A modular class to handle extraction from various document types."""

    def __init__(self, file_path: str, original_filename: str):
        self.file_path = file_path
        self.original_filename = original_filename
        self.extension = Path(original_filename).suffix.lower()

    def extract_text(self) -> str:
        """
        Classifies the document by its extension and calls the appropriate
        extraction method.
        """
        logger.info(f"Classifying '{self.original_filename}' with extension '{self.extension}' for text extraction.")
        
        if self.extension in TEXT_EXTENSIONS:
            return self._read_text()
        elif self.extension in JSON_EXTENSIONS:
            return self._read_json()
        elif self.extension in CSV_EXTENSIONS:
            return self._read_csv()
        elif self.extension in EXCEL_EXTENSIONS:
            return self._read_excel()
        elif self.extension in IMAGE_EXTENSIONS:
            return self._ocr_image()
        elif self.extension in DOCX_EXTENSIONS:
            return self._read_docx()
        elif self.extension in PDF_EXTENSIONS:
            return self._read_pdf_with_ocr_fallback()
        else:
            logger.warning(f"Unsupported file type '{self.extension}' for file '{self.original_filename}'.")
            raise ValueError(f"Unsupported file type: {self.extension}")

    def _read_text(self) -> str:
        """Reads plain text files."""
        logger.debug(f"Reading plain text from '{self.original_filename}'")
        with open(self.file_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()

    def _read_json(self) -> str:
        """Reads and pretty-prints JSON files."""
        logger.debug(f"Reading JSON from '{self.original_filename}'")
        with open(self.file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            # Convert JSON object to a pretty-printed string for context
            return json.dumps(data, indent=2)

    def _read_csv(self) -> str:
        """Reads and formats CSV files as a string."""
        logger.debug(f"Reading CSV from '{self.original_filename}'")
        with open(self.file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            # Convert CSV to a simple string representation
            return "\n".join([", ".join(row) for row in reader])

    def _read_excel(self) -> str:
        """Reads and formats XLSX files as a string."""
        logger.debug(f"Reading Excel (XLSX) from '{self.original_filename}'")
        workbook = openpyxl.load_workbook(self.file_path)
        full_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f"--- Sheet: {sheet_name} ---\n")
            for row in sheet.iter_rows(values_only=True):
                # Join cells, converting non-string types to string and filtering None
                row_text = ", ".join([str(cell) for cell in row if cell is not None])
                if row_text: # Only add non-empty rows
                    full_text.append(row_text)
        return "\n".join(full_text)

    def _ocr_image(self) -> str:
        """Extracts text from an image file using Tesseract OCR."""
        logger.debug(f"Performing OCR on image '{self.original_filename}'")
        try:
            return pytesseract.image_to_string(Image.open(self.file_path))
        except Exception as e:
            logger.error(f"Error during OCR for image '{self.original_filename}': {e}")
            if "Tesseract is not installed" in str(e) or "tesseract not found" in str(e).lower():
                raise EnvironmentError("Tesseract OCR engine not found or not in PATH.")
            raise

    def _read_docx(self) -> str:
        """Extracts text from a DOCX file."""
        logger.debug(f"Reading DOCX from '{self.original_filename}'")
        doc = docx.Document(self.file_path)
        return "\n".join([para.text for para in doc.paragraphs])

    def _read_pdf_with_ocr_fallback(self) -> str:
        """
        Extracts text from a PDF file. First tries direct text extraction.
        If the extracted text is minimal, it performs OCR as a fallback.
        """
        logger.debug(f"Reading PDF from '{self.original_filename}', attempting direct text extraction first.")
        text = ""
        try:
            reader = pypdf.PdfReader(self.file_path)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except Exception as e:
            logger.warning(f"pypdf failed to read '{self.original_filename}': {e}. Proceeding directly to OCR.")
            text = "" # Ensure text is empty if pypdf fails

        # Check if direct text extraction was successful
        if len(text.strip()) >= MIN_PDF_TEXT_LENGTH:
            logger.info(f"Successfully extracted {len(text)} chars directly from PDF '{self.original_filename}'.")
            return text
        
        # If not, perform OCR as a fallback
        logger.warning(f"Minimal text ({len(text.strip())} chars) extracted directly. Attempting OCR fallback for '{self.original_filename}'.")
        try:
            images = convert_from_path(self.file_path)
            ocr_text = "".join(pytesseract.image_to_string(img) for img in images)
            logger.info(f"OCR fallback extracted {len(ocr_text)} chars from PDF '{self.original_filename}'.")
            return ocr_text
        except Exception as ocr_error:
            logger.error(f"OCR fallback failed for '{self.original_filename}': {ocr_error}")
            # If both direct extraction and OCR fail, but direct had *some* text, return that.
            if text.strip():
                logger.warning("Returning minimal text from direct extraction as OCR failed.")
                return text
            # Otherwise, raise an error indicating total failure.
            raise ValueError("Both direct text extraction and OCR failed for the PDF.")